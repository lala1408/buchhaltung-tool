# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import cgi
import html
import json
import mimetypes
import re
import secrets
import shutil
import sys
from copy import copy as copy_style
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote, unquote, urlparse

from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter
from reportlab.lib.colors import HexColor
from reportlab.pdfgen import canvas


APP_DIR = Path(__file__).resolve().parent
SESSIONS_DIR = APP_DIR / "work" / "sessions"
OUTPUTS_DIR = APP_DIR / "outputs"
BACKUPS_DIR = APP_DIR / "backups"
CONFIG_PATH = APP_DIR / "config.json"
DEFAULT_WORKBOOK_PATH = Path(r"C:\Users\Lars\Downloads\SFK-Buchhaltung ab 2022.xlsx")

OUT_COLS = {
    "nr": 9,
    "text": 10,
    "payment": 11,
    "voucher": 12,
    "date": 13,
    "amount": 14,
    "category": 15,
    "note": 17,
}


@dataclass
class WorkbookSuggestion:
    sheet_name: str
    year: int
    next_no: int
    voucher: str
    target_row: int
    existing_text: str
    existing_payment: str
    existing_date: str
    existing_amount: str
    existing_category: str
    existing_note: str
    warning: str


@dataclass
class PdfSuggestion:
    amount: str
    date_value: str
    description: str
    notes: list[str]
    extracted_preview: str


def ensure_dirs() -> None:
    SESSIONS_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)


def load_config() -> dict[str, str]:
    if CONFIG_PATH.exists():
        try:
            data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return {str(key): str(value) for key, value in data.items()}
        except json.JSONDecodeError:
            pass
    if DEFAULT_WORKBOOK_PATH.exists():
        return {"workbook_path": str(DEFAULT_WORKBOOK_PATH)}
    return {"workbook_path": ""}


def save_config(config: dict[str, str]) -> None:
    CONFIG_PATH.write_text(json.dumps(config, indent=2, ensure_ascii=False), encoding="utf-8")


def resolve_workbook_path(path_text: str) -> Path:
    workbook_path = Path(path_text.strip().strip('"'))
    if not workbook_path.exists():
        raise ValueError(f"Excel-Datei nicht gefunden: {workbook_path}")
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("Die Buchhaltungsdatei muss eine .xlsx-Datei sein.")
    return workbook_path.resolve()


def backup_workbook(workbook_path: Path, voucher: str) -> Path:
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_name = sanitize_filename(f"{workbook_path.stem} Backup vor {voucher} {timestamp}.xlsx")
    backup_path = BACKUPS_DIR / backup_name
    shutil.copy2(workbook_path, backup_path)
    return backup_path


def sanitize_filename(name: str, fallback: str = "datei") -> str:
    name = Path(name or fallback).name.strip()
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name)
    name = re.sub(r"\s+", " ", name).strip(" .")
    return name or fallback


def unique_token() -> str:
    return secrets.token_urlsafe(16)


def is_within(child: Path, parent: Path) -> bool:
    try:
        child.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def html_escape(value: Any) -> str:
    return html.escape("" if value is None else str(value), quote=True)


def parse_year_from_sheet(sheet_name: str) -> int | None:
    match = re.search(r"(20\d{2})", sheet_name)
    return int(match.group(1)) if match else None


def find_year_sheets(workbook_path: Path) -> list[int]:
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    years = []
    for name in wb.sheetnames:
        year = parse_year_from_sheet(name)
        if year is not None:
            years.append(year)
    return sorted(set(years), reverse=True)


def sheet_for_year(workbook_path: Path, year: int) -> str:
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    wanted = f"Abrechnung {year}"
    if wanted in wb.sheetnames:
        return wanted
    for name in wb.sheetnames:
        if parse_year_from_sheet(name) == year:
            return name
    raise ValueError(f"Kein Tabellenblatt fuer {year} gefunden.")


def read_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def suggest_from_workbook(workbook_path: Path, year: int | None = None) -> WorkbookSuggestion:
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    years = []
    for name in wb.sheetnames:
        parsed = parse_year_from_sheet(name)
        if parsed is not None:
            years.append(parsed)
    if not years:
        raise ValueError("Keine Jahresblaetter wie 'Abrechnung 2026' gefunden.")
    selected_year = year or max(years)
    sheet_name = f"Abrechnung {selected_year}"
    if sheet_name not in wb.sheetnames:
        matches = [name for name in wb.sheetnames if parse_year_from_sheet(name) == selected_year]
        if not matches:
            raise ValueError(f"Kein Tabellenblatt fuer {selected_year} gefunden.")
        sheet_name = matches[0]

    ws = wb[sheet_name]
    last_voucher_no = 0
    voucher_pattern = re.compile(rf"^A-{selected_year}-(\d+)$", re.IGNORECASE)
    for row in range(4, ws.max_row + 1):
        value = ws.cell(row, OUT_COLS["voucher"]).value
        if not value:
            continue
        match = voucher_pattern.match(str(value).strip())
        if match:
            last_voucher_no = max(last_voucher_no, int(match.group(1)))

    next_no = last_voucher_no + 1
    target_row = 0
    for row in range(4, ws.max_row + 1):
        nr = ws.cell(row, OUT_COLS["nr"]).value
        if isinstance(nr, str) and nr.isdigit():
            nr = int(nr)
        if nr == next_no:
            target_row = row
            break

    warning = ""
    if not target_row:
        for row in range(4, ws.max_row + 1):
            has_summary_formula = any(
                isinstance(ws.cell(row, col).value, str) and ws.cell(row, col).value.startswith("=")
                for col in range(OUT_COLS["nr"], OUT_COLS["note"] + 1)
            )
            if has_summary_formula:
                break
            row_values = [ws.cell(row, col).value for col in OUT_COLS.values()]
            if all(value in (None, "") for value in row_values):
                target_row = row
                break
    if not target_row:
        target_row = ws.max_row + 1
        warning = "Es wurde keine vorbereitete freie Zeile gefunden; die App wuerde am Tabellenende schreiben."

    voucher = f"A-{selected_year}-{next_no:02d}"
    return WorkbookSuggestion(
        sheet_name=sheet_name,
        year=selected_year,
        next_no=next_no,
        voucher=voucher,
        target_row=target_row,
        existing_text=read_cell_text(ws.cell(target_row, OUT_COLS["text"]).value),
        existing_payment=read_cell_text(ws.cell(target_row, OUT_COLS["payment"]).value) or "Überweisung",
        existing_date=read_cell_text(ws.cell(target_row, OUT_COLS["date"]).value),
        existing_amount=read_cell_text(ws.cell(target_row, OUT_COLS["amount"]).value),
        existing_category=read_cell_text(ws.cell(target_row, OUT_COLS["category"]).value),
        existing_note=read_cell_text(ws.cell(target_row, OUT_COLS["note"]).value),
        warning=warning,
    )


def extract_pdf_text(pdf_path: Path) -> str:
    pieces: list[str] = []
    try:
        with pdf_path.open("rb") as handle:
            reader = PdfReader(handle)
            for page in reader.pages:
                pieces.append(page.extract_text() or "")
    except Exception as exc:  # PDF extraction can fail on malformed scans.
        pieces.append(f"[PDF-Text konnte nicht gelesen werden: {exc}]")
    return "\n".join(pieces)


def parse_decimal(value: str) -> Decimal:
    cleaned = str(value).strip()
    cleaned = cleaned.replace("€", "").replace("EUR", "")
    cleaned = cleaned.replace(" ", "")
    if "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    try:
        return Decimal(cleaned)
    except InvalidOperation as exc:
        raise ValueError(f"Betrag ist ungueltig: {value}") from exc


def format_decimal(value: Decimal | str | float | int) -> str:
    decimal_value = value if isinstance(value, Decimal) else Decimal(str(value))
    return f"{decimal_value:.2f}".replace(".", ",")


def guess_amount(text: str) -> str:
    amount_pattern = r"(\d{1,4}(?:\.\d{3})*,\d{2}|\d+\.\d{2})"
    direct_amounts: list[Decimal] = []
    for direct_pattern in (
        rf"(?:gesamtbetrag|rechnungsbetrag)\D{{0,30}}{amount_pattern}",
        rf"{amount_pattern}\D{{0,30}}(?:gesamtbetrag|rechnungsbetrag)",
    ):
        for match in re.finditer(direct_pattern, text, re.IGNORECASE):
            raw = match.group(1)
            try:
                amount = parse_decimal(raw)
            except ValueError:
                continue
            if Decimal("0") < amount <= Decimal("100000"):
                direct_amounts.append(amount)
    if direct_amounts:
        return format_decimal(max(direct_amounts))

    candidates: dict[Decimal, dict[str, Any]] = {}
    pattern = re.compile(rf"(?<![\d./]){amount_pattern}(?![\d./])")
    for match in pattern.finditer(text):
        raw = match.group(1)
        try:
            amount = parse_decimal(raw)
        except ValueError:
            continue
        if amount <= 0 or amount > Decimal("100000"):
            continue
        window = text[max(0, match.start() - 80) : min(len(text), match.end() + 80)].lower()
        score = 0
        for keyword in ("gesamtbetrag", "rechnungsbetrag"):
            if keyword in window:
                score += 8
        for keyword in ("betrag", "ueberweisung", "überweisung", "gesamt", "bezahlt"):
            if keyword in window:
                score += 2
        for keyword in ("ust", "mwst", "umsatzsteuer", "steuer", "netto", "einzel"):
            if keyword in window:
                score -= 5
        entry = candidates.setdefault(amount, {"score": 0, "count": 0})
        entry["score"] += score
        entry["count"] += 1
    if not candidates:
        return ""
    ranked = sorted(
        candidates.items(),
        key=lambda item: (item[1]["score"], item[1]["count"], item[0]),
        reverse=True,
    )
    return format_decimal(ranked[0][0])


def guess_date(text: str) -> str:
    patterns = [
        (re.compile(r"\b(\d{2})[.](\d{2})[.](20\d{2})\b"), "%d.%m.%Y"),
        (re.compile(r"\b(20\d{2})-(\d{2})-(\d{2})\b"), "%Y-%m-%d"),
    ]
    matches: list[tuple[int, date]] = []
    for regex, fmt in patterns:
        for match in regex.finditer(text):
            raw = match.group(0)
            try:
                parsed = datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
            window = text[max(0, match.start() - 80) : min(len(text), match.end() + 80)].lower()
            score = 0
            for keyword in ("ausfuehrung", "ausführung", "buchung", "ueberweisung", "überweisung", "datum"):
                if keyword in window:
                    score += 2
            if "rechnung" in window:
                score += 1
            matches.append((score, parsed))
    if not matches:
        return ""
    matches.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return matches[0][1].isoformat()


def parse_date_value(raw: str) -> str:
    raw = raw.strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


def find_labeled_date(text: str, labels: list[str]) -> str:
    compact = " ".join(text.split())
    date_pattern = r"(\d{2}\.\d{2}\.20\d{2}|20\d{2}-\d{2}-\d{2})"
    for label in labels:
        pattern = rf"{label}[^\d]{{0,80}}{date_pattern}"
        match = re.search(pattern, compact, re.IGNORECASE)
        if match:
            parsed = parse_date_value(match.group(1))
            if parsed:
                return parsed
    return ""


def guess_transfer_booking_date(transfer_text: str) -> str:
    value_date = find_labeled_date(
        transfer_text,
        [
            r"Wertstellung\s*\(Valuta\)",
            r"Wertstellung",
            r"Valutatag",
            r"Valuta",
        ],
    )
    if value_date:
        return value_date
    return find_labeled_date(transfer_text, [r"Buchungstag", r"Buchungsdatum"])


def guess_description(text: str, proof_filename: str) -> str:
    normalized = " ".join(text.split())
    item_patterns = [
        r"\b1\s+(.{5,100}?)\s+\d+\s+(?:Stueck|Stück|Stk\.?)\s+\d",
        r"Bezeichnung\s+Menge.*?\b1\s+(.{5,100}?)\s+\d+\s+",
    ]
    for pattern in item_patterns:
        match = re.search(pattern, normalized, re.IGNORECASE)
        if match:
            item = re.sub(r"\s+", " ", match.group(1)).strip()
            item = re.sub(r"\b(weiss|weiß|blau|schwarz)\b", "", item, flags=re.IGNORECASE).strip()
            if item:
                return item[:90]
    stem = Path(proof_filename).stem
    stem = re.sub(r"^A-\d{4}-\d+\s*", "", stem)
    stem = re.sub(r"\s+\d+[,.]\d{2}\s*(?:€|EUR)?$", "", stem).strip()
    return stem[:90] or "Ausgabe laut Beleg"


def analyze_pdfs(transfer_pdf: Path, proof_pdf: Path) -> PdfSuggestion:
    transfer_text = extract_pdf_text(transfer_pdf)
    proof_text = extract_pdf_text(proof_pdf)
    combined_text = f"{transfer_text}\n{proof_text}"
    amount = guess_amount(combined_text)
    date_value = guess_transfer_booking_date(transfer_text)
    description = guess_description(combined_text, proof_pdf.name)
    notes = []
    if len(" ".join(transfer_text.split())) < 20:
        notes.append("Der Ueberweisungsbeleg liefert kaum lesbaren Text. Datum/Betrag bitte pruefen.")
    if not amount:
        notes.append("Kein Betrag erkannt.")
    if not date_value:
        notes.append("Kein Wertstellungstag oder Buchungstag im Ueberweisungsbeleg erkannt.")
    preview = " ".join(combined_text.split())[:1200]
    return PdfSuggestion(amount, date_value, description, notes, preview)


def copy_row_style(ws: Any, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy_style(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy_style(source.alignment)
        if source.font:
            target.font = copy_style(source.font)
        if source.fill:
            target.fill = copy_style(source.fill)
        if source.border:
            target.border = copy_style(source.border)


def update_workbook(
    workbook_path: Path,
    output_path: Path,
    sheet_name: str,
    row: int,
    next_no: int,
    voucher: str,
    description: str,
    payment: str,
    booking_date: str,
    amount: str,
    category: str,
    note: str,
) -> None:
    wb = load_workbook(workbook_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Tabellenblatt nicht gefunden: {sheet_name}")
    ws = wb[sheet_name]
    if row > ws.max_row:
        copy_row_style(ws, max(4, ws.max_row), row)

    date_value = datetime.strptime(booking_date, "%Y-%m-%d").date()
    amount_value = float(parse_decimal(amount))
    category_value: int | str
    try:
        category_value = int(str(category).strip())
    except ValueError:
        category_value = str(category).strip()

    ws.cell(row, OUT_COLS["nr"]).value = next_no
    ws.cell(row, OUT_COLS["text"]).value = description.strip()
    ws.cell(row, OUT_COLS["payment"]).value = payment.strip() or "Überweisung"
    ws.cell(row, OUT_COLS["voucher"]).value = voucher.strip()
    ws.cell(row, OUT_COLS["date"]).value = date_value
    ws.cell(row, OUT_COLS["amount"]).value = amount_value
    ws.cell(row, OUT_COLS["category"]).value = category_value
    ws.cell(row, OUT_COLS["note"]).value = note.strip() or None

    ws.cell(row, OUT_COLS["date"]).number_format = "DD.MM.YYYY"
    ws.cell(row, OUT_COLS["amount"]).number_format = '#,##0.00'

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_stamp_overlay(width: float, height: float, voucher: str) -> PdfReader:
    packet = BytesIO()
    page_width = float(width)
    page_height = float(height)
    c = canvas.Canvas(packet, pagesize=(page_width, page_height))
    font_size = 16
    text_width = c.stringWidth(voucher, "Helvetica", font_size)
    cx = page_width / 2
    cy = page_height - 58
    rx = max(68, (text_width / 2) + 30)
    ry = 24
    red = HexColor("#ef1b1b")
    c.setStrokeColor(red)
    c.setFillColor(red)
    c.setLineWidth(1.2)
    c.ellipse(cx - rx, cy - ry, cx + rx, cy + ry, stroke=1, fill=0)
    c.setFont("Helvetica", font_size)
    c.drawCentredString(cx, cy - (font_size / 3), voucher)
    c.save()
    packet.seek(0)
    return PdfReader(packet)


def add_pdf_pages(writer: PdfWriter, path: Path, voucher: str | None = None) -> None:
    with path.open("rb") as handle:
        reader = PdfReader(handle)
        for page in reader.pages:
            if voucher:
                width = page.mediabox.width
                height = page.mediabox.height
                overlay_reader = build_stamp_overlay(width, height, voucher)
                page.merge_page(overlay_reader.pages[0])
            writer.add_page(page)


def merge_and_stamp_pdfs(transfer_pdf: Path, proof_pdf: Path, output_path: Path, voucher: str) -> None:
    writer = PdfWriter()
    add_pdf_pages(writer, transfer_pdf, voucher=voucher)
    add_pdf_pages(writer, proof_pdf, voucher=None)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("wb") as handle:
        writer.write(handle)


def make_output_basename(voucher: str, description: str, amount: str) -> str:
    clean_description = sanitize_filename(description, "Beleg")
    clean_description = re.sub(r"\s+", " ", clean_description)[:90].strip()
    clean_amount = amount.strip().replace(".", ",")
    return sanitize_filename(f"{voucher} {clean_description} {clean_amount} EUR")


def render_page(body: str, title: str = "SFK Buchhaltung") -> bytes:
    document = f"""<!doctype html>
<html lang="de">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html_escape(title)}</title>
  <style>
    :root {{
      --bg: #f5f6f2;
      --panel: #ffffff;
      --ink: #17201c;
      --muted: #5c665f;
      --line: #d8ded6;
      --accent: #176b5b;
      --accent-dark: #0f4d42;
      --warn: #a14d16;
      --soft: #edf3ef;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: Arial, Helvetica, sans-serif;
      background: var(--bg);
      color: var(--ink);
      line-height: 1.45;
    }}
    header {{
      border-bottom: 1px solid var(--line);
      background: #fbfcfa;
    }}
    .top {{
      max-width: 1180px;
      margin: 0 auto;
      padding: 18px 24px;
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
    }}
    h1 {{
      margin: 0;
      font-size: 24px;
      letter-spacing: 0;
    }}
    main {{
      max-width: 1180px;
      margin: 0 auto;
      padding: 24px;
    }}
    .grid {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) minmax(320px, 420px);
      gap: 18px;
      align-items: start;
    }}
    section {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 18px;
    }}
    h2 {{
      margin: 0 0 14px;
      font-size: 18px;
      letter-spacing: 0;
    }}
    h3 {{
      margin: 18px 0 10px;
      font-size: 15px;
      letter-spacing: 0;
    }}
    label {{
      display: block;
      font-size: 13px;
      font-weight: 700;
      color: var(--muted);
      margin: 12px 0 6px;
    }}
    input, select, textarea {{
      width: 100%;
      font: inherit;
      border: 1px solid #c8d0ca;
      border-radius: 6px;
      padding: 10px 11px;
      background: #fff;
      color: var(--ink);
    }}
    input[type="file"] {{
      padding: 8px;
      background: #fbfcfb;
    }}
    textarea {{
      min-height: 84px;
      resize: vertical;
    }}
    .row {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 12px;
    }}
    .row-three {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 12px;
    }}
    button, .button {{
      appearance: none;
      border: 0;
      border-radius: 6px;
      padding: 11px 15px;
      background: var(--accent);
      color: white;
      font: inherit;
      font-weight: 700;
      cursor: pointer;
      text-decoration: none;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 8px;
      min-height: 42px;
    }}
    button:hover, .button:hover {{ background: var(--accent-dark); }}
    .secondary {{
      background: #e8ece8;
      color: var(--ink);
    }}
    .secondary:hover {{ background: #dfe5df; }}
    .actions {{
      display: flex;
      gap: 10px;
      flex-wrap: wrap;
      margin-top: 16px;
    }}
    .meta {{
      color: var(--muted);
      font-size: 13px;
    }}
    .notice {{
      border-left: 4px solid var(--warn);
      background: #fff7ed;
      padding: 10px 12px;
      border-radius: 4px;
      color: #623413;
      margin: 12px 0;
      font-size: 14px;
    }}
    .ok {{
      border-left-color: var(--accent);
      background: var(--soft);
      color: #163c34;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 14px;
    }}
    th, td {{
      text-align: left;
      border-bottom: 1px solid var(--line);
      padding: 8px 6px;
      vertical-align: top;
    }}
    th {{
      color: var(--muted);
      font-size: 12px;
      text-transform: uppercase;
      letter-spacing: 0;
    }}
    code {{
      background: #eef1ee;
      border-radius: 4px;
      padding: 2px 5px;
      font-family: Consolas, "Courier New", monospace;
      font-size: 0.95em;
    }}
    pre {{
      max-height: 260px;
      overflow: auto;
      background: #f1f4f1;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 12px;
      white-space: pre-wrap;
      font-size: 12px;
    }}
    @media (max-width: 860px) {{
      .grid, .row, .row-three {{ grid-template-columns: 1fr; }}
      .top {{ align-items: flex-start; flex-direction: column; }}
      main {{ padding: 16px; }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="top">
      <h1>SFK Buchhaltung</h1>
      <div class="meta">Ausgabenbelege automatisieren</div>
    </div>
  </header>
  <main>
    {body}
  </main>
</body>
</html>"""
    return document.encode("utf-8")


def upload_form(error: str = "") -> bytes:
    year_value = datetime.now().year
    config = load_config()
    workbook_path = config.get("workbook_path", "")
    workbook_status = "bereit" if workbook_path and Path(workbook_path).exists() else "nicht gefunden"
    error_html = f'<div class="notice">{html_escape(error)}</div>' if error else ""
    body = f"""
<div class="grid">
  <section>
    <h2>Beleg vorbereiten</h2>
    {error_html}
    <form method="post" action="/analyze" enctype="multipart/form-data">
      <div class="row">
        <div>
          <label for="year">Jahr</label>
          <input id="year" name="year" type="number" min="2022" max="2099" value="{year_value}" required>
        </div>
        <div>
          <label for="payment">Zahlungsart</label>
          <select id="payment" name="payment">
            <option selected>Überweisung</option>
            <option>Lastschrift</option>
            <option>Bar</option>
            <option>Kartenzahlung</option>
          </select>
        </div>
      </div>
      <label for="workbook_path">Buchhaltungs-Excel</label>
      <input id="workbook_path" name="workbook_path" value="{html_escape(workbook_path)}" required>
      <label for="transfer_pdf">Überweisungsbeleg</label>
      <input id="transfer_pdf" name="transfer_pdf" type="file" accept=".pdf" required>
      <label for="proof_pdf">Rechnung oder Nachweis</label>
      <input id="proof_pdf" name="proof_pdf" type="file" accept=".pdf" required>
      <div class="actions">
        <button type="submit">Analysieren</button>
      </div>
    </form>
  </section>
  <section>
    <h2>Nummernkreis</h2>
    <table>
      <tr><th>Excel</th><td>{html_escape(workbook_status)}</td></tr>
      <tr><th>Ausgaben</th><td><code>A-JAHR-NR</code></td></tr>
      <tr><th>Excel-Blatt</th><td><code>Abrechnung JAHR</code></td></tr>
      <tr><th>PDF-Reihenfolge</th><td>Überweisung, danach Nachweis</td></tr>
      <tr><th>Stempel</th><td>Rot eingekreist nur auf dem Überweisungsbeleg</td></tr>
      <tr><th>Ausgabe</th><td>Excel-Update, Backup und zusammengefügter PDF-Beleg</td></tr>
    </table>
  </section>
</div>
"""
    return render_page(body)


def analysis_page(
    token: str,
    workbook_name: str,
    workbook_path: Path,
    transfer_name: str,
    proof_name: str,
    wb_suggestion: WorkbookSuggestion,
    pdf_suggestion: PdfSuggestion,
    payment_fallback: str,
) -> bytes:
    description = wb_suggestion.existing_text or pdf_suggestion.description
    amount = wb_suggestion.existing_amount or pdf_suggestion.amount
    date_value = wb_suggestion.existing_date or pdf_suggestion.date_value
    category = wb_suggestion.existing_category or "3"
    payment = wb_suggestion.existing_payment or payment_fallback or "Überweisung"
    warning_lines = []
    if wb_suggestion.warning:
        warning_lines.append(wb_suggestion.warning)
    warning_lines.extend(pdf_suggestion.notes)
    warnings = "".join(f'<div class="notice">{html_escape(line)}</div>' for line in warning_lines)

    existing_hint = ""
    if wb_suggestion.existing_text or wb_suggestion.existing_amount:
        existing_hint = '<div class="notice ok">Die App nutzt die bereits vorbereitete Zeile in der Excel als Vorschlag.</div>'

    body = f"""
<div class="grid">
  <section>
    <h2>Vorschlag prüfen</h2>
    {warnings}
    {existing_hint}
    <form method="post" action="/create">
      <input type="hidden" name="token" value="{html_escape(token)}">
      <input type="hidden" name="workbook_name" value="{html_escape(workbook_name)}">
      <input type="hidden" name="workbook_path" value="{html_escape(str(workbook_path))}">
      <input type="hidden" name="transfer_name" value="{html_escape(transfer_name)}">
      <input type="hidden" name="proof_name" value="{html_escape(proof_name)}">
      <input type="hidden" name="sheet_name" value="{html_escape(wb_suggestion.sheet_name)}">
      <input type="hidden" name="target_row" value="{wb_suggestion.target_row}">
      <input type="hidden" name="next_no" value="{wb_suggestion.next_no}">
      <div class="row-three">
        <div>
          <label for="voucher">Beleg</label>
          <input id="voucher" name="voucher" value="{html_escape(wb_suggestion.voucher)}" required>
        </div>
        <div>
          <label for="booking_date">Datum</label>
          <input id="booking_date" name="booking_date" type="date" value="{html_escape(date_value)}" required>
        </div>
        <div>
          <label for="amount">Betrag</label>
          <input id="amount" name="amount" value="{html_escape(format_decimal(parse_decimal(amount)) if amount else '')}" required>
        </div>
      </div>
      <label for="description">Buchungstext</label>
      <input id="description" name="description" value="{html_escape(description)}" required>
      <div class="row-three">
        <div>
          <label for="payment">Zahlungsart</label>
          <select id="payment" name="payment">
            {select_options(["Überweisung", "Lastschrift", "Bar", "Kartenzahlung"], payment)}
          </select>
        </div>
        <div>
          <label for="category">Kategorie</label>
          <input id="category" name="category" value="{html_escape(category)}" required>
        </div>
        <div>
          <label for="note">Bemerkung</label>
          <input id="note" name="note" value="{html_escape(wb_suggestion.existing_note)}">
        </div>
      </div>
      <label>
        <input type="checkbox" name="update_source" value="1" checked style="width:auto; margin-right:8px;">
        Excel-Datei direkt aktualisieren und vorher Backup anlegen
      </label>
      <div class="actions">
        <button type="submit">Beleg erzeugen</button>
        <a class="button secondary" href="/">Neu starten</a>
      </div>
    </form>
  </section>
  <section>
    <h2>Analyse</h2>
    <table>
      <tr><th>Excel</th><td>{html_escape(str(workbook_path))}</td></tr>
      <tr><th>Blatt</th><td>{html_escape(wb_suggestion.sheet_name)}</td></tr>
      <tr><th>Zeile</th><td>{wb_suggestion.target_row}</td></tr>
      <tr><th>Nächste Nr.</th><td>{wb_suggestion.next_no}</td></tr>
      <tr><th>Überweisung</th><td>{html_escape(transfer_name)}</td></tr>
      <tr><th>Nachweis</th><td>{html_escape(proof_name)}</td></tr>
    </table>
    <h3>PDF-Textauszug</h3>
    <pre>{html_escape(pdf_suggestion.extracted_preview or "Kein lesbarer Text gefunden.")}</pre>
  </section>
</div>
"""
    return render_page(body)


def select_options(options: list[str], selected: str) -> str:
    html_options = []
    for option in options:
        attr = " selected" if option == selected else ""
        html_options.append(f'<option{attr}>{html_escape(option)}</option>')
    if selected and selected not in options:
        html_options.append(f'<option selected>{html_escape(selected)}</option>')
    return "\n".join(html_options)


def result_page(
    voucher: str,
    excel_link: str,
    pdf_link: str,
    output_dir: Path,
    updated_workbook: Path | None,
    backup_path: Path | None,
) -> bytes:
    update_rows = ""
    if updated_workbook is not None:
        update_rows += f'<tr><th>Excel aktualisiert</th><td><code>{html_escape(str(updated_workbook))}</code></td></tr>'
    if backup_path is not None:
        update_rows += f'<tr><th>Backup</th><td><code>{html_escape(str(backup_path))}</code></td></tr>'
    body = f"""
<section>
  <h2>Fertig: {html_escape(voucher)}</h2>
  <div class="notice ok">Excel und PDF wurden erzeugt.</div>
  <table>
    <tr><th>Ordner</th><td><code>{html_escape(str(output_dir))}</code></td></tr>
    {update_rows}
  </table>
  <div class="actions">
    <a class="button" href="{html_escape(excel_link)}">Excel herunterladen</a>
    <a class="button" href="{html_escape(pdf_link)}">PDF herunterladen</a>
    <a class="button secondary" href="/">Nächster Beleg</a>
  </div>
</section>
"""
    return render_page(body)


class AccountingRequestHandler(BaseHTTPRequestHandler):
    server_version = "SFKBuchhaltung/0.1"

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/":
            self.respond_html(upload_form())
            return
        if parsed.path.startswith("/download/"):
            self.serve_download(parsed.path)
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Nicht gefunden")

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/analyze":
                self.handle_analyze()
                return
            if parsed.path == "/create":
                self.handle_create()
                return
            self.send_error(HTTPStatus.NOT_FOUND, "Nicht gefunden")
        except Exception as exc:
            self.respond_html(upload_form(str(exc)), status=HTTPStatus.BAD_REQUEST)

    def respond_html(self, content: bytes, status: HTTPStatus = HTTPStatus.OK) -> None:
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def parse_form(self) -> cgi.FieldStorage:
        return cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type", ""),
            },
            encoding="utf-8",
            errors="replace",
        )

    def get_text_field(self, form: cgi.FieldStorage, name: str, default: str = "") -> str:
        field = form[name] if name in form else None
        if field is None:
            return default
        if isinstance(field, list):
            field = field[0]
        return str(field.value or default).strip()

    def save_upload(self, form: cgi.FieldStorage, name: str, folder: Path, fallback: str) -> Path:
        field = form[name] if name in form else None
        if field is None or isinstance(field, list) or not getattr(field, "file", None):
            raise ValueError(f"Upload fehlt: {name}")
        filename = sanitize_filename(field.filename or fallback, fallback)
        path = folder / filename
        with path.open("wb") as output:
            shutil.copyfileobj(field.file, output)
        if path.stat().st_size == 0:
            raise ValueError(f"Upload ist leer: {filename}")
        return path

    def handle_analyze(self) -> None:
        ensure_dirs()
        form = self.parse_form()
        token = unique_token()
        session_dir = SESSIONS_DIR / token
        session_dir.mkdir(parents=True, exist_ok=True)

        workbook_path = resolve_workbook_path(self.get_text_field(form, "workbook_path"))
        save_config({"workbook_path": str(workbook_path)})
        transfer_pdf = self.save_upload(form, "transfer_pdf", session_dir, "ueberweisung.pdf")
        proof_pdf = self.save_upload(form, "proof_pdf", session_dir, "nachweis.pdf")
        year_text = self.get_text_field(form, "year")
        year = int(year_text) if year_text else None
        payment = self.get_text_field(form, "payment", "Überweisung")

        wb_suggestion = suggest_from_workbook(workbook_path, year)
        pdf_suggestion = analyze_pdfs(transfer_pdf, proof_pdf)
        self.respond_html(
            analysis_page(
                token=token,
                workbook_name=workbook_path.name,
                workbook_path=workbook_path,
                transfer_name=transfer_pdf.name,
                proof_name=proof_pdf.name,
                wb_suggestion=wb_suggestion,
                pdf_suggestion=pdf_suggestion,
                payment_fallback=payment,
            )
        )

    def handle_create(self) -> None:
        ensure_dirs()
        form = self.parse_form()
        token = self.get_text_field(form, "token")
        if not re.fullmatch(r"[A-Za-z0-9_-]+", token or ""):
            raise ValueError("Ungueltiger Sitzungsschluessel.")
        session_dir = SESSIONS_DIR / token
        if not is_within(session_dir, SESSIONS_DIR) or not session_dir.exists():
            raise ValueError("Sitzung nicht gefunden.")

        workbook_path = resolve_workbook_path(self.get_text_field(form, "workbook_path"))
        save_config({"workbook_path": str(workbook_path)})

        transfer_name = sanitize_filename(self.get_text_field(form, "transfer_name"), "ueberweisung.pdf")
        proof_name = sanitize_filename(self.get_text_field(form, "proof_name"), "nachweis.pdf")
        transfer_pdf = session_dir / transfer_name
        proof_pdf = session_dir / proof_name
        if not transfer_pdf.exists() or not proof_pdf.exists():
            pdf_files = sorted(session_dir.glob("*.pdf"))
            if len(pdf_files) < 2:
                raise ValueError("Zwei PDF-Dateien wurden nicht gefunden.")
            transfer_pdf = pdf_files[0]
            proof_pdf = pdf_files[1]

        sheet_name = self.get_text_field(form, "sheet_name")
        target_row = int(self.get_text_field(form, "target_row"))
        next_no = int(self.get_text_field(form, "next_no"))
        voucher = self.get_text_field(form, "voucher")
        description = self.get_text_field(form, "description")
        payment = self.get_text_field(form, "payment", "Überweisung")
        booking_date = self.get_text_field(form, "booking_date")
        amount = self.get_text_field(form, "amount")
        category = self.get_text_field(form, "category")
        note = self.get_text_field(form, "note")
        update_source = self.get_text_field(form, "update_source") == "1"

        output_dir = OUTPUTS_DIR / token
        basename = make_output_basename(voucher, description, amount)
        excel_output = output_dir / f"{basename}.xlsx"
        pdf_output = output_dir / f"{basename}.pdf"

        update_workbook(
            workbook_path=workbook_path,
            output_path=excel_output,
            sheet_name=sheet_name,
            row=target_row,
            next_no=next_no,
            voucher=voucher,
            description=description,
            payment=payment,
            booking_date=booking_date,
            amount=amount,
            category=category,
            note=note,
        )
        backup_path = None
        updated_workbook = None
        if update_source:
            backup_path = backup_workbook(workbook_path, voucher)
            shutil.copy2(excel_output, workbook_path)
            updated_workbook = workbook_path

        merge_and_stamp_pdfs(transfer_pdf, proof_pdf, pdf_output, voucher)

        excel_link = f"/download/{quote(token)}/{quote(excel_output.name)}"
        pdf_link = f"/download/{quote(token)}/{quote(pdf_output.name)}"
        self.respond_html(result_page(voucher, excel_link, pdf_link, output_dir, updated_workbook, backup_path))

    def serve_download(self, path: str) -> None:
        parts = [unquote(part) for part in path.split("/") if part]
        if len(parts) != 3:
            self.send_error(HTTPStatus.NOT_FOUND, "Nicht gefunden")
            return
        _, token, filename = parts
        if not re.fullmatch(r"[A-Za-z0-9_-]+", token or ""):
            self.send_error(HTTPStatus.FORBIDDEN, "Ungueltiger Pfad")
            return
        file_path = OUTPUTS_DIR / token / sanitize_filename(filename)
        if not is_within(file_path, OUTPUTS_DIR) or not file_path.exists():
            self.send_error(HTTPStatus.NOT_FOUND, "Datei nicht gefunden")
            return
        content_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
        data = file_path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        ascii_name = file_path.name.encode("ascii", "ignore").decode("ascii") or "download"
        encoded_name = quote(file_path.name)
        self.send_header("Content-Disposition", f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{encoded_name}')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, format: str, *args: Any) -> None:
        sys.stdout.write("%s - %s\n" % (self.address_string(), format % args))


def main() -> None:
    parser = argparse.ArgumentParser(description="Lokale Buchhaltungs-App fuer SFK-Belege")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", default=8501, type=int)
    args = parser.parse_args()
    ensure_dirs()
    server = ThreadingHTTPServer((args.host, args.port), AccountingRequestHandler)
    print(f"SFK Buchhaltung laeuft auf http://{args.host}:{args.port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServer beendet.")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
